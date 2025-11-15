#!/usr/bin/env python3
"""
screenplay_parser.py

Сервис парсинга типизированных сценариев в excel-таблицу.
Использует GPU-оптимизированную LLM-архитектуру для извлечения метаданных.

Автор: Production Pipeline Parser
Версия: 2.0.0 (LLM-based)
"""

import argparse
import json
import os
import sys
import gc
from typing import List, Dict, Any, Optional
from dataclasses import dataclass, field
import logging
from pathlib import Path
import warnings
warnings.filterwarnings('ignore')

import pandas as pd
from docx import Document
from tqdm import tqdm
import numpy as np

# Настройка логирования
logging.basicConfig(
    level=logging.INFO, 
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('screenplay_parser.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Импорт нового LLM-движка
try:
    from app.llm_engine import LLMEngine
except ImportError:
    try:
        from llm_engine import LLMEngine
    except ImportError:
        logger.error("Не удалось импортировать LLMEngine. Убедитесь, что llm_engine.py находится в app/")
        LLMEngine = None


# -----------------------------
#  Структуры данных
# -----------------------------

@dataclass
class SceneMetadata:
    """Структура для хранения метаданных сцены"""
    scene_number: str = ""
    episode: str = ""
    scene_type: str = ""
    location: str = ""
    sublocation: str = ""
    time_of_day: str = ""
    synopsis: str = ""
    characters: List[str] = field(default_factory=list)
    extras: str = ""
    extras_count: int = 0
    props: List[str] = field(default_factory=list)
    vehicles: List[str] = field(default_factory=list)
    special_fx: List[str] = field(default_factory=list)
    costumes: List[str] = field(default_factory=list)
    makeup: List[str] = field(default_factory=list)
    stunts: bool = False
    pyrotechnics: bool = False
    special_equipment: List[str] = field(default_factory=list)
    notes: str = ""
    raw_text: str = ""
    confidence_score: float = 0.0


# -----------------------------
#  Парсинг сценария (LLM-based)
# -----------------------------

class ScenarioParser:
    """Класс для парсинга сценария с использованием LLM"""
    
    def __init__(self, config_path: str = "config.yaml", preset: str = "full", custom_entities: Optional[List[str]] = None):
        """
        Инициализация парсера
        
        Args:
            config_path: Путь к конфигурационному файлу
            preset: Пресет для извлечения сущностей ("basic", "extended", "full")
            custom_entities: Кастомный список сущностей для извлечения (если None, используется preset)
        """
        self.config_path = config_path
        self.preset = preset
        self.custom_entities = custom_entities
        self.llm_engine = None
        
        if LLMEngine is None:
            logger.error("LLMEngine недоступен. Установите зависимости: pip install transformers torch")
            raise RuntimeError("LLMEngine недоступен")
        
        try:
            logger.info("Инициализация LLM-движка...")
            self.llm_engine = LLMEngine(config_path=config_path)
            logger.info("LLM-движок успешно инициализирован")
        except Exception as e:
            logger.error(f"Ошибка инициализации LLM-движка: {e}")
            raise
        
        self.scenes = []
    
    def parse_screenplay(self, text: str) -> List[SceneMetadata]:
        """
        Основной метод парсинга сценария
        
        Args:
            text: Текст сценария
            
        Returns:
            Список SceneMetadata с извлеченными данными
        """
        logger.info("Начинаю парсинг сценария через LLM...")
        
        if not self.llm_engine:
            raise RuntimeError("LLM-движок не инициализирован")
        
        try:
            # Используем полный пайплайн LLM-движка
            results = self.llm_engine.process_screenplay(
                screenplay_text=text,
                preset=self.preset,
                custom_entities=self.custom_entities
            )
            
            logger.info(f"LLM обработал {len(results)} сцен")
            
            # Преобразуем результаты в SceneMetadata
            self.scenes = []
            for result in results:
                metadata = self._convert_to_metadata(result)
                self.scenes.append(metadata)
            
            logger.info(f"Успешно обработано {len(self.scenes)} сцен")
            return self.scenes
            
        except Exception as e:
            logger.error(f"Ошибка при парсинге сценария: {e}")
            import traceback
            traceback.print_exc()
            raise
    
    def _convert_to_metadata(self, result: Dict[str, Any]) -> SceneMetadata:
        """Преобразует результат LLM в SceneMetadata"""
        metadata = SceneMetadata()
        
        # Базовые поля
        metadata.scene_number = str(result.get('scene_number', ''))
        metadata.raw_text = result.get('scene_text', '')
        
        # Парсим заголовок сцены для извлечения базовой информации
        scene_title = result.get('scene_title', '')
        self._parse_scene_title(scene_title, metadata)
        
        # Извлекаем сущности из результата LLM
        location = self._safe_get(result, 'location', '')
        # Дополнительная валидация локации - убираем глаголы
        metadata.location = self._clean_location(location, result.get('scene_text', ''))
        metadata.sublocation = self._safe_get(result, 'sublocation', '')
        metadata.time_of_day = self._safe_get(result, 'time_of_day', '')
        metadata.scene_type = self._safe_get(result, 'scene_type', '')
        metadata.synopsis = self._safe_get(result, 'synopsis', '')
        
        # Списки с валидацией
        metadata.characters = self._clean_characters(self._safe_get_list(result, 'characters'))
        metadata.props = self._safe_get_list(result, 'props')
        metadata.vehicles = self._safe_get_list(result, 'vehicles')
        metadata.special_fx = self._safe_get_list(result, 'vfx')
        metadata.costumes = self._clean_costumes(self._safe_get_list(result, 'costumes'))
        metadata.makeup = self._safe_get_list(result, 'makeup')
        metadata.special_equipment = self._clean_equipment(self._safe_get_list(result, 'special_equipment'))
        
        # Массовка
        crowd = self._safe_get(result, 'crowd', '')
        if crowd:
            metadata.extras = crowd
        
        # Количество массовки - сначала из crowd_count, потом из crowd
        crowd_count = self._safe_get(result, 'crowd_count', None)
        if crowd_count is not None:
            try:
                metadata.extras_count = int(crowd_count)
            except (ValueError, TypeError):
                metadata.extras_count = 0
        elif crowd:
            # Пытаемся извлечь число из описания массовки
            import re
            # Ищем числа в тексте массовки
            numbers = re.findall(r'\d+', str(crowd))
            if numbers:
                try:
                    # Берем первое найденное число
                    metadata.extras_count = int(numbers[0])
                except (ValueError, TypeError):
                    metadata.extras_count = 0
            else:
                metadata.extras_count = 0
        else:
            metadata.extras_count = 0
        
        # Булевы значения
        metadata.stunts = self._safe_get(result, 'stunts', False)
        metadata.pyrotechnics = self._safe_get(result, 'pyrotechnics', False)
        
        # Примечания
        metadata.notes = self._safe_get(result, 'notes', '')
        
        # Звуковые эффекты (если есть)
        sfx = self._safe_get_list(result, 'sfx')
        if sfx:
            if not metadata.notes:
                metadata.notes = f"Звуковые эффекты: {', '.join(sfx)}"
            else:
                metadata.notes += f"\nЗвуковые эффекты: {', '.join(sfx)}"
        
        # Уверенность (можно добавить из результата LLM, если есть)
        metadata.confidence_score = result.get('confidence', 0.8)
        
        return metadata
    
    def _parse_scene_title(self, title: str, metadata: SceneMetadata):
        """Парсит заголовок сцены для извлечения базовой информации"""
        if not title:
            return
        
        # Ищем тип сцены
        import re
        scene_type_match = re.search(r'(ИНТ|ЭКСТ|НАТ|INT|EXT)', title, re.IGNORECASE)
        if scene_type_match:
            metadata.scene_type = scene_type_match.group(0).upper()
            if metadata.scene_type in ['INT', 'EXT']:
                metadata.scene_type = 'ИНТ' if metadata.scene_type == 'INT' else 'ЭКСТ'
        
        # Ищем время суток
        time_match = re.search(r'(ДЕНЬ|НОЧЬ|УТРО|ВЕЧЕР|РАССВЕТ|ЗАКАТ)', title, re.IGNORECASE)
        if time_match:
            metadata.time_of_day = time_match.group(0).upper()
        
        # Если локация не была извлечена LLM, пытаемся извлечь из заголовка
        if not metadata.location:
            # Убираем тип сцены и время суток, оставляем локацию
            location_text = re.sub(r'(ИНТ|ЭКСТ|НАТ|INT|EXT)\.?\s*', '', title, flags=re.IGNORECASE)
            location_text = re.sub(r'(ДЕНЬ|НОЧЬ|УТРО|ВЕЧЕР|РАССВЕТ|ЗАКАТ)', '', location_text, flags=re.IGNORECASE)
            location_text = location_text.strip(' -.,')
            if location_text:
                metadata.location = location_text
    
    def _safe_get(self, d: Dict, key: str, default: Any = None) -> Any:
        """Безопасное получение значения из словаря"""
        value = d.get(key, default)
        if value is None:
            return default
        return value
    
    def _safe_get_list(self, d: Dict, key: str) -> List[str]:
        """Безопасное получение списка из словаря"""
        value = d.get(key, [])
        if isinstance(value, list):
            return [str(item) for item in value if item]
        elif isinstance(value, str):
            # Если это строка, пытаемся разбить по запятым
            return [item.strip() for item in value.split(',') if item.strip()]
        return []
    
    def _clean_location(self, location: str, scene_text: str) -> str:
        """Очистка локации от глаголов и действий"""
        if not location:
            return location
        
        import re
        # Паттерны глаголов и действий
        invalid_patterns = [
            r'отделаются', r'плавают', r'идут', r'стоят', r'сидят',
            r'бегут', r'говорят', r'смотрят', r'делают', r'находятся'
        ]
        
        location_lower = location.lower()
        for pattern in invalid_patterns:
            if re.search(pattern, location_lower):
                logger.warning(f"Обнаружена некорректная локация: '{location}'. Очищаю...")
                # Пытаемся найти настоящую локацию в тексте
                location_match = re.search(
                    r'(?:ИНТ|ЭКСТ|НАТ|INT|EXT)\.\s*([А-ЯЁ\w\s]+?)(?:\s*[-–—]\s*|\.|$)',
                    scene_text,
                    re.IGNORECASE
                )
                if location_match:
                    found = location_match.group(1).strip()
                    if found and not any(re.search(p, found.lower()) for p in invalid_patterns):
                        return found
                return ""  # Если не нашли, возвращаем пустую строку
        
        return location
    
    def _clean_characters(self, characters: List[str]) -> List[str]:
        """Очистка персонажей от массовки и глаголов"""
        if not characters:
            return characters
        
        import re
        invalid_words = ['толпа', 'массовка', 'люди', 'прохожие', 'студенты', 'официанты']
        invalid_verbs = [r'плавают', r'идут', r'стоят', r'сидят', r'бегут', r'говорят']
        
        cleaned = []
        for char in characters:
            char_str = str(char).strip()
            if not char_str:
                continue
            
            char_lower = char_str.lower()
            
            # Пропускаем массовку
            if any(word in char_lower for word in invalid_words):
                continue
            
            # Пропускаем глаголы
            if any(re.search(verb, char_lower) for verb in invalid_verbs):
                continue
            
            # Пропускаем слишком длинные "имена"
            if len(char_str) > 30:
                continue
            
            cleaned.append(char_str)
        
        return cleaned
    
    def _clean_equipment(self, equipment: List[str]) -> List[str]:
        """Очистка спецоборудования - только профессиональное съемочное оборудование"""
        if not equipment:
            return equipment
        
        import re
        valid_keywords = ['кран', 'дрон', 'стабилизатор', 'тележка', 'журавль', 'микрофон',
                         'освещение', 'камера', 'оператор', 'съемочн', 'техническ',
                         'подвес', 'трос', 'подъемник', 'платформа', 'рельс']
        invalid_patterns = [r'стол', r'стул', r'кровать', r'диван', r'телефон', r'компьютер',
                           r'книга', r'бумага', r'ручка', r'одежда', r'костюм', r'грим']
        
        cleaned = []
        for item in equipment:
            item_str = str(item).strip().lower()
            if not item_str:
                continue
            
            # Пропускаем обычный реквизит
            if any(re.search(pattern, item_str) for pattern in invalid_patterns):
                logger.warning(f"Удалено из спецоборудования (это реквизит): '{item}'")
                continue
            
            # Проверяем, что это профессиональное оборудование
            if any(keyword in item_str for keyword in valid_keywords):
                cleaned.append(item)
            else:
                logger.warning(f"Удалено из спецоборудования (не профессиональное): '{item}'")
        
        return cleaned
    
    def _clean_costumes(self, costumes: List[str]) -> List[str]:
        """Очистка костюмов - только конкретные описания"""
        if not costumes:
            return costumes
        
        costume_keywords = ['костюм', 'платье', 'одежда', 'форма', 'униформа', 'наряд']
        cleaned = []
        
        for costume in costumes:
            costume_str = str(costume).strip().lower()
            if not costume_str:
                continue
            
            # Проверяем, что это действительно описание костюма
            if any(keyword in costume_str for keyword in costume_keywords):
                cleaned.append(costume)
            elif len(costume_str) > 5 and costume_str not in ['одежда', 'костюм', 'форма']:
                cleaned.append(costume)
        
        return cleaned


# -----------------------------
#  Чтение файлов
# -----------------------------

def read_docx(path: str) -> str:
    """Читает .docx файл"""
    try:
        doc = Document(path)
        paragraphs = []
        
        for para in doc.paragraphs:
            text = para.text.strip()
            if text:
                if getattr(para.style, 'name', '').startswith('Heading'):
                    text = f"\n\n{text}\n"
                paragraphs.append(text)
        
        return "\n".join(paragraphs)
        
    except Exception as e:
        logger.error(f"Ошибка чтения файла {path}: {e}")
        raise


def read_pdf(path: str) -> str:
    """Читает .pdf файл"""
    try:
        import pdfplumber
        text_parts = []
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
        return "\n\n".join(text_parts)
    except ImportError:
        raise RuntimeError("pdfplumber не установлен. Установите: pip install pdfplumber")
    except Exception as e:
        logger.error(f"Ошибка чтения PDF {path}: {e}")
        raise


# -----------------------------
#  Экспорт в Excel
# -----------------------------

def create_production_table(scenes: List[SceneMetadata]) -> pd.DataFrame:
    """Создает таблицу для КПП из списка сцен"""
    rows = []
    
    for scene in scenes:
        row = {
            "Серия": scene.episode or "01",
            "Сцена": scene.scene_number,
            "Режим": scene.time_of_day,
            "Инт/Нат": scene.scene_type,
            "Объект": scene.location,
            "Подобъект": scene.sublocation,
            "Синопсис": scene.synopsis[:200] if scene.synopsis else "",
            "Персонажи": ", ".join(scene.characters[:8]) if scene.characters else "",
            "Массовка": scene.extras,
            "Кол-во массовки": scene.extras_count if scene.extras_count else "",
            "Реквизит": ", ".join(scene.props[:8]) if scene.props else "",
            "Игровой транспорт": ", ".join(scene.vehicles) if scene.vehicles else "",
            "Художники": "",
            "Грим": ", ".join(scene.makeup) if scene.makeup else "",
            "Костюм": ", ".join(scene.costumes) if scene.costumes else "",
            "Каскадеры": "Да" if scene.stunts else "",
            "Пиротехника": "Да" if scene.pyrotechnics else "",
            "Спец. оборудование": ", ".join(scene.special_equipment) if scene.special_equipment else "",
            "Примечание": scene.notes,
            "Уверенность": f"{scene.confidence_score:.0%}" if scene.confidence_score > 0 else ""
        }
        rows.append(row)
    
    df = pd.DataFrame(rows)
    
    # Сортировка по номеру сцены
    try:
        df['scene_num'] = df['Сцена'].astype(str).str.extract(r'(\d+)').astype(float)
        df = df.sort_values('scene_num').drop('scene_num', axis=1)
    except:
        pass
    
    return df


def export_to_excel(df: pd.DataFrame, output_path: str):
    """Экспортирует DataFrame в Excel с форматированием"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        ws = wb.active
        ws.title = "КПП"
        
        ws.append([f"КПП - Календарно-постановочный план"])
        ws.merge_cells('A1:T1')
        
        header_font = Font(bold=True, size=14)
        ws['A1'].font = header_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.append([])
        
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[3]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        column_widths = {
            'A': 8, 'B': 10, 'C': 10, 'D': 12, 'E': 25, 'F': 25, 'G': 40,
            'H': 30, 'I': 20, 'J': 10, 'K': 30, 'L': 20, 'M': 20, 'N': 20,
            'O': 20, 'P': 12, 'Q': 12, 'R': 25, 'S': 30, 'T': 12
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Колонка "Уверенность" (T = 20)
        confidence_col = 20
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=confidence_col, max_col=confidence_col):
            for cell in row:
                if cell.value:
                    try:
                        confidence = float(str(cell.value).strip('%')) / 100
                        if confidence >= 0.8:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif confidence >= 0.6:
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    except:
                        pass
        
        wb.save(output_path)
        logger.info(f"Таблица успешно экспортирована в {output_path}")
        
    except Exception as e:
        logger.error(f"Ошибка при экспорте в Excel: {e}")
        df.to_excel(output_path, index=False, sheet_name='КПП')
        logger.info(f"Таблица экспортирована в упрощенном формате")


# -----------------------------
#  Статистика и отчеты
# -----------------------------

def print_statistics(scenes: List[SceneMetadata], df: pd.DataFrame):
    """Выводит статистику по обработанным сценам"""
    print("\n" + "="*70)
    print(" " * 20 + "СТАТИСТИКА ПАРСИНГА")
    print("="*70)
    
    print(f"\n📊 ОСНОВНЫЕ ПОКАЗАТЕЛИ:")
    print(f"  • Всего сцен: {len(scenes)}")
    print(f"  • Интерьеры: {sum(1 for s in scenes if s.scene_type in ['INT', 'ИНТ'])}")
    print(f"  • Натура: {sum(1 for s in scenes if s.scene_type in ['EXT', 'ЭКСТ', 'НАТ'])}")
    print(f"  • Дневные сцены: {sum(1 for s in scenes if 'ДЕНЬ' in s.time_of_day.upper())}")
    print(f"  • Ночные сцены: {sum(1 for s in scenes if 'НОЧЬ' in s.time_of_day.upper())}")
    
    locations = df['Объект'].value_counts() if 'Объект' in df.columns else pd.Series()
    print(f"\n📍 ЛОКАЦИИ:")
    print(f"  • Уникальных локаций: {len(locations)}")
    print(f"  • Топ-5 локаций:")
    for loc, count in locations.head(5).items():
        print(f"    - {loc}: {count} сцен")
    
    all_characters = []
    for scene in scenes:
        all_characters.extend(scene.characters)
    unique_chars = list(set(all_characters))
    
    print(f"\n👥 ПЕРСОНАЖИ:")
    print(f"  • Уникальных персонажей: {len(unique_chars)}")
    if unique_chars:
        from collections import Counter
        char_counts = Counter(all_characters)
        print(f"  • Топ-5 персонажей:")
        for char, count in char_counts.most_common(5):
            print(f"    - {char}: {count} сцен")
    
    print(f"\n🎬 ПРОИЗВОДСТВЕННЫЕ ТРЕБОВАНИЯ:")
    print(f"  • Сцены с массовкой: {sum(1 for s in scenes if s.extras)}")
    print(f"  • Сцены с транспортом: {sum(1 for s in scenes if s.vehicles)}")
    print(f"  • Сцены со спецэффектами: {sum(1 for s in scenes if s.special_fx)}")
    print(f"  • Сцены с трюками: {sum(1 for s in scenes if s.stunts)}")
    print(f"  • Сцены с пиротехникой: {sum(1 for s in scenes if s.pyrotechnics)}")
    
    if any(s.confidence_score > 0 for s in scenes):
        avg_confidence = np.mean([s.confidence_score for s in scenes if s.confidence_score > 0])
        print(f"\n📈 КАЧЕСТВО ПАРСИНГА:")
        print(f"  • Средняя уверенность: {avg_confidence:.0%}")
        print(f"  • Сцены с высокой уверенностью (>80%): {sum(1 for s in scenes if s.confidence_score > 0.8)}")
        print(f"  • Сцены с низкой уверенностью (<60%): {sum(1 for s in scenes if 0 < s.confidence_score < 0.6)}")
    
    print("\n" + "="*70)


# -----------------------------
#  CLI
# -----------------------------

def main():
    parser = argparse.ArgumentParser(
        description="🎬 Парсер сценариев для создания КПП (LLM-based)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Примеры использования:
  python screenplay_parser.py -i scenario.docx -o production.xlsx
  python screenplay_parser.py -i scenario.docx -o production.xlsx --preset basic
  python screenplay_parser.py -i scenario.docx -o production.xlsx --preset extended
  python screenplay_parser.py -i scenario.docx -o production.xlsx --config custom_config.yaml
        """
    )
    
    parser.add_argument(
        "--input", "-i",
        required=True,
        help="Путь к файлу сценария (.docx или .pdf)"
    )
    parser.add_argument(
        "--output", "-o",
        default="production_table.xlsx",
        help="Путь для сохранения Excel таблицы"
    )
    parser.add_argument(
        "--config", "-c",
        default="config.yaml",
        help="Путь к конфигурационному файлу (по умолчанию: config.yaml)"
    )
    parser.add_argument(
        "--preset", "-p",
        choices=["basic", "extended", "full"],
        default="full",
        help="Пресет для извлечения сущностей (по умолчанию: full)"
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Режим отладки с подробным выводом"
    )
    
    args = parser.parse_args()
    
    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)
    
    if not os.path.exists(args.input):
        logger.error(f"❌ Файл не найден: {args.input}")
        sys.exit(1)
    
    file_ext = os.path.splitext(args.input)[1].lower()
    if file_ext not in ['.docx', '.pdf']:
        logger.error("❌ Поддерживаются только форматы .docx и .pdf")
        sys.exit(1)
    
    try:
        logger.info(f"📖 Чтение файла {args.input}...")
        if file_ext == '.pdf':
            text = read_pdf(args.input)
        else:
            text = read_docx(args.input)
        logger.info(f"✓ Прочитано {len(text)} символов")
        
        logger.info(f"🤖 Инициализация LLM-парсера (пресет: {args.preset})...")
        parser_obj = ScenarioParser(
            config_path=args.config,
            preset=args.preset
        )
        
        logger.info("🔄 Начинаю обработку сценария через LLM...")
        scenes = parser_obj.parse_screenplay(text)
        logger.info(f"✓ Обработано сцен: {len(scenes)}")
        
        logger.info("📊 Создание таблицы production...")
        df = create_production_table(scenes)
        
        logger.info(f"💾 Сохранение в {args.output}...")
        export_to_excel(df, args.output)
        
        print_statistics(scenes, df)
        
        print(f"\n✅ ГОТОВО! Файл сохранен: {args.output}")
        print(f"📂 Откройте файл в Excel для просмотра и редактирования")
        
    except KeyboardInterrupt:
        logger.info("\n⚠️ Прервано пользователем")
        sys.exit(0)
    except Exception as e:
        logger.error(f"❌ Ошибка: {e}")
        if args.debug:
            import traceback
            traceback.print_exc()
        sys.exit(1)
    finally:
        gc.collect()


if __name__ == "__main__":
    main()
